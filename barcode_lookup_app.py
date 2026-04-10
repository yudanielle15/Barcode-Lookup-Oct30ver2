import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook

st.set_page_config(page_title="Biomarker Sample Barcode Lookup Web App", layout="centered")
st.title("🔬 Biomarker Sample Barcode Lookup Web App")
st.write("Upload your Excel file locally, and scan or enter a barcode.")

# Initialize session state
if "df" not in st.session_state:
    st.session_state.df = None
if "barcode_input" not in st.session_state:
    st.session_state.barcode_input = ""

# To track barcode input in session state
barcode_input_placeholder = st.empty()

uploaded_file = st.file_uploader("📁 Upload your sample Excel file", type=["xlsx"])

if uploaded_file:
    try:
        # Load the file if it's not already loaded
        if st.session_state.df is None:
            df = pd.read_excel(uploaded_file)
            if "Scan_Status" not in df.columns:
                df["Scan_Status"] = ""
            st.session_state.df = df
        st.success("✅ File loaded. Ready to scan.")

        # --- Display the loaded table ---
        st.subheader("📋 Loaded Table")
        st.dataframe(st.session_state.df)

        # --- Barcode input ---
        barcode_input = barcode_input_placeholder.text_input("🧪 Scan or type barcode:", value=st.session_state.barcode_input)

        if barcode_input:
            df = st.session_state.df
            current_match = df[df['Barcode'].astype(str) == str(barcode_input)]

            if current_match.empty:
                # Update error message to include the scanned barcode
                st.error(f"❌ No match found for {barcode_input}.")
            else:
                st.success("✅ Sample found:")
                # Update Scan_Status in backend
                df.loc[df['Barcode'].astype(str) == str(barcode_input), 'Scan_Status'] = "Matched"
                st.session_state.df = df
                st.info(f"🗸 Scan status updated for barcode: {barcode_input}")

                # --- Highlight and Show current match (highlight columns in yellow) ---
                st.subheader("🔹 Current Match(es)")

                # Apply styling to highlight specific columns in matched rows
                def highlight_row(row):
                    # Apply yellow background for "Screen ID", "Visit", "Sample Name"
                    styles = [''] * len(row)
                    highlight_cols = ['Screen ID', 'Visit', 'Sample Name']
                    for i, col in enumerate(row.index):
                        if col in highlight_cols:
                            styles[i] = 'background-color: yellow'
                    return styles

                # Apply the highlighting function to the matched rows
                styled_match = current_match.style.apply(highlight_row, axis=1)

                # Display the styled dataframe
                st.dataframe(styled_match)

                # --- Full table with highlighting for current match ---
                st.subheader("📋 Full Table")

                # Apply styling only to the matched rows in the full table
                def highlight_full_table(row):
                    # Check if this row is a match and highlight the relevant columns
                    styles = [''] * len(row)
                    highlight_cols = ['Screen ID', 'Visit', 'Sample Name']
                    if row['Barcode'] == barcode_input:
                        for i, col in enumerate(row.index):
                            if col in highlight_cols:
                                styles[i] = 'background-color: yellow'
                    return styles

                # Apply the highlight to the entire table
                styled_full_table = df.style.apply(highlight_full_table, axis=1)

                # Display the full table with highlighted columns for the current match
                st.dataframe(styled_full_table)

                # --- Download updated Excel ---
                if st.session_state.df is not None:
                    original_filename = uploaded_file.name
                    new_filename = original_filename.replace(".xlsx", "_Scanned.xlsx")
                    uploaded_file.seek(0)
                    wb = load_workbook(uploaded_file)
                    ws = wb.active

                    # Add Scan_Status column if missing
                    if "Scan_Status" not in [cell.value for cell in ws[1]]:
                        ws.cell(row=1, column=ws.max_column + 1, value="Scan_Status")

                    # Map headers
                    header = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

                    df = st.session_state.df
                    for i, val in enumerate(df['Scan_Status'], start=2):
                        ws.cell(row=i, column=header["Scan_Status"], value=val)

                    buffer = BytesIO()
                    wb.save(buffer)
                    buffer.seek(0)

                    st.download_button(
                        label="💾 Download Updated Excel File",
                        data=buffer,
                        file_name=new_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

        # --- Clear the barcode input UI after processing ---
        st.session_state.barcode_input = ""  # Reset the barcode input value in session state
        barcode_input_placeholder.empty()  # Clear the input UI field

        # Re-render barcode input placeholder with an empty value
        barcode_input_placeholder.text_input("🧪 Scan or type barcode:", value="", key="barcode_input")

    except Exception as e:
        st.error(f"❌ Error reading file: {e}")
else:
    st.info("⬆️ Please upload an Excel file to begin.")
